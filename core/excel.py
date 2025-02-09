import openpyxl
from openpyxl.styles import Alignment, Font

from datetime import datetime, timedelta

from core.utils import MessageFormatter


class ExcelExporter:
    @staticmethod
    def filter_records_by_date(records: list, days: int):
        # Get the current date and subtract the number of days specified
        cutoff_date = datetime.now() - timedelta(days=days)

        # Convert cutoff date to string for comparison (assuming format is "YYYY-MM-DD HH:MM:SS")
        filtered_records = [
            record for record in records if datetime.strptime(record["created_at"], "%Y-%m-%d %H:%M:%S") > cutoff_date
        ]
        return filtered_records

    @staticmethod
    async def export_to_excel(records: list, output_file: str, chat_id: int, days: int = 0) -> str:
        # Filter records based on the number of days
        if days > 0:
            records = ExcelExporter.filter_records_by_date(records, days)

        # Create a workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transaction History"

        # Set up the header
        ws.append(["Date and Time", "Transaction Description"])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Process the records
        for record in records:
            date_time = record.get("created_at", "")
            amount = record.get("amount", 0)
            currency = record.get("currency", "").upper()
            is_vidano = record.get("is_vidano", False)

            # Format the transaction description
            if is_vidano:
                description = f"Issued: {amount} {currency} U Joker Trade"
            else:
                description = f"Added: {amount} {currency}"

            ws.append([date_time, description])

        # Calculate total profit and add it at the end
        vidano_balance = await MessageFormatter.get_total_balance(records, chat_id)
        ws.append(["Total volume:", f"{round(vidano_balance, 2)} USD"])
        ws.cell(ws.max_row, 2).font = Font(bold=True)

        # Adjust column width and alignment
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save the file
        wb.save(output_file)
        print(f"Excel file saved as {output_file}")

        return output_file
